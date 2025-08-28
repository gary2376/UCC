#!/usr/bin/env python
# -*- coding: utf-8 -*-
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Avg, Count
from django.utils.decorators import method_decorator
from django.views import View
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from datetime import datetime, timedelta
from typing import Any
import json
import hashlib
import pandas as pd
import os
import re
import math
import io
import calendar

from app.models.models import GreenBeanInboundRecord, RawMaterialWarehouseRecord, RawMaterialMonthlySummary, UserActivityLog, FileUploadRecord, UploadRecordRelation
from app.serializers.user_serializer import (
    GreenBeanInboundRecordSerializer,
    RawMaterialWarehouseRecordSerializer,
    RawMaterialMonthlySummarySerializer
)
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from app.utils.activity_logger import log_user_activity, get_recent_user_activities, get_important_user_activities, _log_to_django_admin
from django.db import transaction
from django.core.files.storage import default_storage
from app.utils.permission_utils import get_user_accessible_sections, require_green_bean_permission, require_raw_material_permission


class ERPDashboardView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """ERP 系統儀表板 - 只需要登入"""
    permission_required = ('app.view_greenbeaninboundrecord', 'app.view_rawmaterialwarehouserecord')
    raise_exception = True
    template_name = 'erp/dashboard.html'
    
    def has_permission(self):
        user = self.request.user
        # 允許 superuser 或有生豆入庫權限或有原料倉管理權限的用戶訪問
        return (user.is_superuser or 
                user.has_perm('app.view_greenbeaninboundrecord') or 
                user.has_perm('app.view_rawmaterialwarehouserecord'))
    
    def get(self, request):
        # 使用 Django 預設的權限檢查機制，會自動返回 403 Forbidden
        if not self.has_permission():
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        # 獲取使用者權限
        user_permissions = get_user_accessible_sections(request.user)
        
        # 根據權限過濾統計數據
        stats = {}
        
        if user_permissions['green_bean']:
            # 今日入庫記錄數量
            today = datetime.now().date()
            today_records = GreenBeanInboundRecord.objects.filter(
                record_time__date=today
            ).count()
            
            # 總生豆入庫量（所有時間的總重量）
            total_weight = GreenBeanInboundRecord.objects.filter(
                measured_weight_kg__isnull=False
            ).aggregate(total=Sum('measured_weight_kg'))['total'] or 0
            
            # 本週總重量
            week_start = today - timedelta(days=today.weekday())
            week_weight = GreenBeanInboundRecord.objects.filter(
                record_time__date__gte=week_start,
                measured_weight_kg__isnull=False
            ).aggregate(total=Sum('measured_weight_kg'))['total'] or 0
            
            # 異常記錄數量
            abnormal_count = GreenBeanInboundRecord.objects.filter(is_abnormal=True).count()
            
            # 本月平均重量
            current_month_avg = GreenBeanInboundRecord.objects.filter(
                record_time__month=datetime.now().month,
                record_time__year=datetime.now().year,
                measured_weight_kg__isnull=False
            ).aggregate(avg=Avg('measured_weight_kg'))['avg'] or 0
            
            stats.update({
                'total_green_bean_records': GreenBeanInboundRecord.objects.count(),
                'recent_abnormal_records': abnormal_count,
                'current_month_records': GreenBeanInboundRecord.objects.filter(
                    record_time__month=datetime.now().month,
                    record_time__year=datetime.now().year
                ).count(),
                'today_green_bean_records': today_records,
                'total_green_bean_weight': round(float(total_weight), 2),
                'week_total_weight': round(float(week_weight), 2),
                'month_avg_weight': round(float(current_month_avg), 2)
            })
        
        if user_permissions['raw_material']:
            # 低庫存商品數量
            low_inventory_threshold = 100
            low_inventory_count = RawMaterialWarehouseRecord.objects.filter(
                current_inventory__lt=low_inventory_threshold,
                current_inventory__gt=0
            ).count()
            
            # 總庫存價值（假設有單價欄位，這裡用庫存數量代替）
            total_inventory = RawMaterialWarehouseRecord.objects.aggregate(
                total=Sum('current_inventory')
            )['total'] or 0
            
            # 本月出入庫次數
            current_month_transactions = RawMaterialWarehouseRecord.objects.filter(
                record_time__month=datetime.now().month,
                record_time__year=datetime.now().year
            ).count()
            
            stats.update({
                'total_raw_material_records': RawMaterialWarehouseRecord.objects.count(),
                'current_month_raw_material_records': current_month_transactions,
                'low_inventory_count': low_inventory_count,
                'total_inventory_amount': total_inventory
            })
        
        # 根據權限獲取最近的記錄
        recent_green_bean_records = []
        recent_raw_material_records = []
        low_inventory_items = []
        
        if user_permissions['green_bean']:
            recent_green_bean_records = GreenBeanInboundRecord.objects.order_by('-record_time')[:10]
        
        if user_permissions['raw_material']:
            recent_raw_material_records = RawMaterialWarehouseRecord.objects.order_by('-created_at')[:10]
            # 庫存警示（低庫存商品）
            low_inventory_threshold = 100
            low_inventory_items = RawMaterialWarehouseRecord.objects.filter(
                current_inventory__lt=low_inventory_threshold,
                current_inventory__gt=0
            ).order_by('current_inventory')[:10]
        
        # 獲取圖表數據
        from app.utils.activity_logger import get_weekly_charts_data
        import json
        weekly_charts_data = json.dumps(get_weekly_charts_data())
        
        context = {
            'stats': stats,
            'recent_green_bean_records': recent_green_bean_records,
            'recent_raw_material_records': recent_raw_material_records,
            'low_inventory_items': low_inventory_items,
            'weekly_charts_data': weekly_charts_data,
            'user_permissions': user_permissions,
        }
        
        return render(request, self.template_name, context)


class ERPCleanDashboardView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """ERP 純淨儀表板 (無側邊選單) - 只需要登入"""
    permission_required = ('app.view_greenbeaninboundrecord', 'app.view_rawmaterialwarehouserecord')
    raise_exception = True
    template_name = 'erp/dashboard_clean.html'
    
    def has_permission(self):
        user = self.request.user
        # 允許 superuser 或有生豆入庫權限或有原料倉管理權限的用戶訪問
        return (user.is_superuser or 
                user.has_perm('app.view_greenbeaninboundrecord') or 
                user.has_perm('app.view_rawmaterialwarehouserecord'))
    
    def get(self, request):
        # 使用 Django 預設的權限檢查機制，會自動返回 403 Forbidden
        if not self.has_permission():
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

        # 不記錄查看儀表板的活動，因為用戶不想看到 view 操作

        # 獲取使用者權限
        user_permissions = get_user_accessible_sections(request.user)

        # 根據權限過濾統計數據
        stats = {}

        if user_permissions['green_bean']:
            # 今日入庫記錄數量
            today = datetime.now().date()
            today_records = GreenBeanInboundRecord.objects.filter(
                record_time__date=today
            ).count()
            
            # 總生豆入庫量（所有時間的總重量）
            total_weight = GreenBeanInboundRecord.objects.filter(
                measured_weight_kg__isnull=False
            ).aggregate(total=Sum('measured_weight_kg'))['total'] or 0
            
            # 本週總重量
            week_start = today - timedelta(days=today.weekday())
            week_weight = GreenBeanInboundRecord.objects.filter(
                record_time__date__gte=week_start,
                measured_weight_kg__isnull=False
            ).aggregate(total=Sum('measured_weight_kg'))['total'] or 0
            
            # 異常記錄數量
            abnormal_count = GreenBeanInboundRecord.objects.filter(is_abnormal=True).count()
            
            # 本月平均重量
            current_month_avg = GreenBeanInboundRecord.objects.filter(
                record_time__month=datetime.now().month,
                record_time__year=datetime.now().year,
                measured_weight_kg__isnull=False
            ).aggregate(avg=Avg('measured_weight_kg'))['avg'] or 0
            
            stats.update({
                'total_green_bean_records': GreenBeanInboundRecord.objects.count(),
                'recent_abnormal_records': abnormal_count,
                'current_month_records': GreenBeanInboundRecord.objects.filter(
                    record_time__month=datetime.now().month,
                    record_time__year=datetime.now().year
                ).count(),
                'today_green_bean_records': today_records,
                'total_green_bean_weight': round(float(total_weight), 2),
                'week_total_weight': round(float(week_weight), 2),
                'month_avg_weight': round(float(current_month_avg), 2)
            })

        if user_permissions['raw_material']:
            # 低庫存商品數量
            low_inventory_threshold = 100
            low_inventory_count = RawMaterialWarehouseRecord.objects.filter(
                current_inventory__lt=low_inventory_threshold,
                current_inventory__gt=0
            ).count()
            
            # 總庫存數量
            total_inventory = RawMaterialWarehouseRecord.objects.aggregate(
                total=Sum('current_inventory')
            )['total'] or 0
            
            # 本月出入庫次數
            current_month_transactions = RawMaterialWarehouseRecord.objects.filter(
                created_at__month=datetime.now().month,
                created_at__year=datetime.now().year
            ).count()
            
            stats.update({
                'total_raw_material_records': RawMaterialWarehouseRecord.objects.count(),
                'current_month_raw_material_records': current_month_transactions,
                'low_inventory_count': low_inventory_count,
                'total_inventory_amount': total_inventory
            })

        # 根據權限獲取記錄
        recent_records = []
        low_inventory_items = []

        if user_permissions['green_bean']:
            recent_records = GreenBeanInboundRecord.objects.order_by('-record_time')[:50]

        if user_permissions['raw_material']:
            # 庫存警示（低庫存商品）
            low_inventory_threshold = 100  # 可以設定為設置項
            low_inventory_items = RawMaterialWarehouseRecord.objects.filter(
                current_inventory__lt=low_inventory_threshold,
                current_inventory__gt=0
            ).order_by('current_inventory')[:10]

        # 最近的重要用戶活動記錄（包含所有檔案和資料操作）
        recent_activities = get_important_user_activities(limit=30)

        # 獲取本週記錄統計
        from app.utils.activity_logger import get_weekly_records_comparison, get_weekly_charts_data
        import json
        weekly_comparison = get_weekly_records_comparison()
        weekly_charts_data = json.dumps(get_weekly_charts_data())

        context = {
            'stats': stats,
            'recent_records': recent_records,
            'low_inventory_items': low_inventory_items,
            'recent_activities': recent_activities,
            'weekly_comparison': weekly_comparison,
            'weekly_charts_data': weekly_charts_data,
            'user_permissions': user_permissions,
        }

        return render(request, self.template_name, context)


@api_view(['GET'])
@require_green_bean_permission('view')
def green_bean_records_api(request):
    """生豆入庫記錄 API - 需要ERP查看權限"""
    try:
        # 查詢參數
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        search = request.GET.get('search', '')
        order_number = request.GET.get('order_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        is_abnormal = request.GET.get('is_abnormal', '')
        
        # 基本查詢
        queryset = GreenBeanInboundRecord.objects.all()
        
        # 應用過濾條件
        if search:
            queryset = queryset.filter(
                Q(green_bean_name__icontains=search) |
                Q(green_bean_code__icontains=search) |
                Q(order_number__icontains=search)
            )
        
        if order_number:
            queryset = queryset.filter(order_number__icontains=order_number)
        
        if start_date:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            queryset = queryset.filter(record_time__gte=start_datetime)
        
        if end_date:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            queryset = queryset.filter(record_time__lt=end_datetime)
        
        if is_abnormal:
            queryset = queryset.filter(is_abnormal=(is_abnormal.lower() == 'true'))
        
        # 排序
        queryset = queryset.order_by('-record_time')
        
        # 分頁
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)
        
        # 序列化數據
        serializer = GreenBeanInboundRecordSerializer(page_obj.object_list, many=True)
        
        return Response({
            'success': True,
            'data': serializer.data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_count': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous()
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@require_raw_material_permission('view')
def raw_material_records_api(request):
    """原料倉記錄 API - 需要ERP查看權限"""
    try:
        # 查詢參數
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        search = request.GET.get('search', '')
        product_code = request.GET.get('product_code', '')
        low_inventory = request.GET.get('low_inventory', '')
        
        # 基本查詢
        queryset = RawMaterialWarehouseRecord.objects.all()
        
        # 應用過濾條件
        if search:
            queryset = queryset.filter(
                Q(product_name__icontains=search) |
                Q(product_code__icontains=search) |
                Q(factory_batch_number__icontains=search)
            )
        
        if product_code:
            queryset = queryset.filter(product_code__icontains=product_code)
        
        if low_inventory:
            threshold = int(low_inventory)
            queryset = queryset.filter(
                current_inventory__lt=threshold,
                current_inventory__gt=0
            )
        
        # 排序
        queryset = queryset.order_by('-created_at')
        
        # 分頁
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)
        
        # 序列化數據
        serializer = RawMaterialWarehouseRecordSerializer(page_obj.object_list, many=True)
        
        return Response({
            'success': True,
            'data': serializer.data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_count': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous()
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def inventory_statistics_api(request):
    """庫存統計 API - 需要ERP查看權限"""
    try:
        # 總庫存統計
        total_stats = RawMaterialWarehouseRecord.objects.aggregate(
            total_inventory=Sum('current_inventory'),
            total_incoming=Sum('incoming_stock'),
            total_outgoing=Sum('outgoing_stock'),
            avg_inventory=Avg('current_inventory')
        )
        
        # 低庫存商品
        low_inventory_threshold = int(request.GET.get('threshold', 100))
        low_inventory_items = RawMaterialWarehouseRecord.objects.filter(
            current_inventory__lt=low_inventory_threshold,
            current_inventory__gt=0
        ).values('product_code', 'product_name', 'current_inventory').order_by('current_inventory')[:20]
        
        # 按產品類別統計
        product_stats = RawMaterialWarehouseRecord.objects.values('product_name').annotate(
            item_count=Count('id'),
            total_inventory=Sum('current_inventory'),
            avg_inventory=Avg('current_inventory')
        ).order_by('-total_inventory')[:10]
        
        return Response({
            'success': True,
            'data': {
                'total_stats': total_stats,
                'low_inventory_items': list(low_inventory_items),
                'product_stats': list(product_stats)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def production_statistics_api(request):
    """生產統計 API - 需要ERP查看權限"""
    try:
        # 時間範圍
        days = int(request.GET.get('days', 30))
        start_date = datetime.now() - timedelta(days=days)
        
        # 生產統計
        production_stats = GreenBeanInboundRecord.objects.filter(
            record_time__gte=start_date
        ).aggregate(
            total_records=Count('id'),
            total_weight=Sum('measured_weight_kg'),
            avg_weight=Avg('measured_weight_kg'),
            abnormal_count=Count('id', filter=Q(is_abnormal=True))
        )
        
        # 每日生產量
        daily_production = GreenBeanInboundRecord.objects.filter(
            record_time__gte=start_date
        ).extra(
            select={'day': 'DATE(record_time)'}
        ).values('day').annotate(
            daily_weight=Sum('measured_weight_kg'),
            daily_count=Count('id')
        ).order_by('day')
        
        # 按生豆類型統計
        bean_type_stats = GreenBeanInboundRecord.objects.filter(
            record_time__gte=start_date
        ).values('green_bean_name').annotate(
            total_weight=Sum('measured_weight_kg'),
            record_count=Count('id')
        ).order_by('-total_weight')[:10]
        
        return Response({
            'success': True,
            'data': {
                'production_stats': production_stats,
                'daily_production': list(daily_production),
                'bean_type_stats': list(bean_type_stats)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@require_green_bean_permission('view')
def green_bean_records_view(request):
    """生豆入庫記錄頁面視圖"""
    # 獲取所有記錄
    records = GreenBeanInboundRecord.objects.all().order_by('-record_time')
    
    # 分頁處理
    paginator = Paginator(records, 20)  # 每頁20條記錄
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 統計資料
    total_records = records.count()
    abnormal_records = records.filter(is_abnormal=True).count()
    current_month_records = records.filter(
        record_time__month=datetime.now().month,
        record_time__year=datetime.now().year
    ).count()
    
    # 最近上傳記錄
    recent_uploads = FileUploadRecord.objects.filter(
        file_type='green_bean'
    ).order_by('-upload_time')[:5]
    
    # 獲取與生豆入庫記錄相關的用戶活動記錄
    from app.utils.activity_logger import get_important_user_activities
    from django.contrib.contenttypes.models import ContentType
    
    # 獲取生豆記錄的 ContentType
    green_bean_content_type = ContentType.objects.get_for_model(GreenBeanInboundRecord)
    
    # 獲取所有重要活動，並過濾生豆相關的活動
    all_activities = get_important_user_activities(limit=100)
    green_bean_activities = []
    
    for activity in all_activities:
        # 包含生豆相關的活動：
        # 1. 直接關聯到 GreenBeanInboundRecord 的活動
        # 2. 描述中包含生豆相關關鍵字的活動
        # 3. 檔案上傳活動（可能包含生豆數據）
        if (activity.content_type == green_bean_content_type or 
            '生豆入庫記錄' in activity.description or 
            activity.action in ['upload', 'delete_upload_record'] or
            (activity.details and any(key in activity.details for key in ['order_number', 'green_bean_name', 'deleted_record']))):
            green_bean_activities.append(activity)
        
        # 限制返回數量
        if len(green_bean_activities) >= 20:
            break
    
    # 獲取生豆名稱列表
    from app.utils.green_bean_utils import get_green_bean_names
    green_bean_names = get_green_bean_names()
    
    context = {
        'page_obj': page_obj,
        'total_records': total_records,
        'abnormal_records': abnormal_records,
        'current_month_records': current_month_records,
        'recent_uploads': recent_uploads,
        'green_bean_activities': green_bean_activities,
        'green_bean_names': green_bean_names,
    }
    
    return render(request, 'erp/green_bean_records.html', context)


@require_green_bean_permission('add')
def green_bean_upload_page(request):
    """生豆入庫記錄上傳頁面"""
    return render(request, 'erp/import_data.html', {
        'title': '生豆入庫記錄上傳',
        'upload_url': '/erp/green-bean-records/upload-file/',
        'records_url': '/erp/green-bean-records/uploads/'
    })


@require_green_bean_permission('add')
@csrf_exempt
@require_http_methods(["POST"])
def green_bean_upload_file(request):
    """生豆入庫記錄檔案上傳處理"""
    try:
        uploaded_file = request.FILES.get('file')
        
        if not uploaded_file:
            return JsonResponse({'success': False, 'message': '請選擇要上傳的檔案'})
        
        # 檢查檔案格式
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': '只支援 .xlsx 和 .xls 格式的檔案'})
        
        # 計算檔案雜湊值
        file_hash = calculate_file_hash(uploaded_file)
        
        # 檢查是否為重複檔案
        existing_file = FileUploadRecord.objects.filter(file_hash=file_hash).first()
        if existing_file:
            return JsonResponse({
                'success': False, 
                'message': f'此檔案已於 {existing_file.upload_time.strftime("%Y-%m-%d %H:%M")} 上傳過',
                'duplicate': True
            })
        
        # 使用事務處理檔案上傳
        with transaction.atomic():
            # 創建上傳記錄
            upload_record = FileUploadRecord.objects.create(
                file_name=uploaded_file.name,
                file_hash=file_hash,
                file_size=uploaded_file.size,
                uploaded_by=request.user,
                file_type='green_bean',
                status='pending'
            )
            
            try:
                # 處理Excel文件
                df = pd.read_excel(uploaded_file)
                
                print(f"原始欄位名稱: {df.columns.tolist()}")
                
                # 清理欄位名稱：移除換行符和空格
                df.columns = df.columns.str.replace('\n', '').str.strip()
                
                print(f"清理後欄位名稱: {df.columns.tolist()}")
                
                # 清理數據：移除完全空白的行
                df = df.dropna(how='all')
                
                # 清理數據：移除關鍵欄位都是空的行
                key_columns = ['單號', '生豆名稱', '生豆料號']
                df = df.dropna(subset=key_columns, how='all')
                
                print(f"清理後剩餘 {len(df)} 行數據")
                
                # 檢查必要欄位 - 根據實際Excel檔案調整
                required_columns = ['單號', '生豆名稱', '生豆料號']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    upload_record.status = 'failed'
                    upload_record.error_message = f'缺少必要欄位: {", ".join(missing_columns)}'
                    upload_record.save()
                    return JsonResponse({
                        'success': False, 
                        'message': f'檔案格式錯誤，缺少必要欄位: {", ".join(missing_columns)}'
                    })
                
                # 定義helper函數
                def safe_numeric(value, default=None):
                    if pd.isna(value) or str(value).strip() in ['', 'nan', 'None']:
                        return default
                    try:
                        result = pd.to_numeric(value, errors='coerce')
                        if pd.isna(result):
                            return default
                        return float(result)  # 確保返回數值類型
                    except:
                        return default
                
                def safe_string(value, default=''):
                    if pd.isna(value) or str(value).strip() in ['nan', 'None']:
                        return default
                    return str(value).strip()
                
                def safe_string_from_numeric(value, default=''):
                    """處理可能是數字的字符串欄位"""
                    if pd.isna(value) or str(value).strip() in ['nan', 'None']:
                        return default
                    # 如果是數字，轉換為整數字符串（去掉小數點）
                    try:
                        if isinstance(value, (int, float)):
                            return str(int(value))
                        return str(value).strip()
                    except:
                        return str(value).strip()
                
                def safe_integer(value, default=None):
                    """安全轉換為整數"""
                    if pd.isna(value) or str(value).strip() in ['', 'nan', 'None']:
                        return default
                    try:
                        result = pd.to_numeric(value, errors='coerce')
                        if pd.isna(result):
                            return default
                        return int(result)
                    except:
                        return default
                
                # 處理資料 - 根據實際Excel欄位對應
                created_records = []
                skipped_rows = 0
                
                for index, row in df.iterrows():
                    try:
                        # 檢查必要欄位是否為空或nan
                        order_number = str(row.get('單號', '')).strip()
                        green_bean_name = str(row.get('生豆名稱', '')).strip()
                        green_bean_code = safe_string_from_numeric(row.get('生豆料號', ''))
                        
                        print(f"處理第 {index + 1} 行: 單號='{order_number}', 生豆名稱='{green_bean_name}', 料號='{green_bean_code}'")
                        
                        # 跳過空行或nan資料
                        if (order_number in ['', 'nan', 'None'] or 
                            green_bean_name in ['', 'nan', 'None'] or 
                            green_bean_code in ['', 'nan', 'None']):
                            print(f"跳過第 {index + 1} 行：空白或無效資料")
                            skipped_rows += 1
                            continue
                        
                        # 處理日期時間欄位
                        record_time = None
                        if pd.notna(row.get('記錄時間')):
                            try:
                                record_time = pd.to_datetime(row.get('記錄時間'))
                            except:
                                record_time = datetime.now()
                        else:
                            record_time = datetime.now()
                        
                        work_start_time = None
                        if pd.notna(row.get('作業開始時間')):
                            try:
                                work_start_time = pd.to_datetime(row.get('作業開始時間'))
                            except:
                                work_start_time = None
                        
                        work_end_time = None
                        if pd.notna(row.get('作業結束時間')):
                            try:
                                work_end_time = pd.to_datetime(row.get('作業結束時間'))
                            except:
                                work_end_time = None
                        
                        # 建立記錄
                        record = GreenBeanInboundRecord.objects.create(
                            # 基本資訊
                            order_number=order_number,
                            roasted_item_sequence=safe_integer(row.get('炒豆項次')),
                            green_bean_item_sequence=safe_integer(row.get('生豆項次')),
                            batch_sequence=safe_integer(row.get('波次')),
                            execution_status=safe_string(row.get('執行狀態')),
                            
                            # 生豆資訊
                            green_bean_code=safe_string_from_numeric(row.get('生豆料號')),
                            green_bean_name=green_bean_name,
                            green_bean_batch_number=safe_string_from_numeric(row.get('生豆批號')),
                            green_bean_storage_silo=safe_string(row.get('生豆入庫筒倉')),
                            
                            # 重量和數量
                            bag_weight_kg=safe_numeric(row.get('一袋重量(kg)')),
                            input_bag_count=safe_integer(row.get('投入袋數')),
                            required_weight_kg=safe_numeric(row.get('需求重量(kg)')),
                            measured_weight_kg=safe_numeric(row.get('生豆量測重量(kg)')),
                            manual_input_weight_kg=safe_numeric(row.get('手動投入重量(kg)')),
                            
                            # 時間資訊
                            record_time=record_time,
                            work_start_time=work_start_time,
                            work_end_time=work_end_time,
                            work_duration=safe_string(row.get('作業時間')),
                            
                            # 其他
                            ico_code=safe_string(row.get('ICO')),
                            remark=safe_string(row.get('備註')),
                            is_abnormal=bool(row.get('異常') == 'Y') if pd.notna(row.get('異常')) else False
                        )
                        created_records.append(record)
                        print(f"成功創建記錄: {record.order_number} - {record.green_bean_name}")
                        
                        # 創建關聯記錄
                        UploadRecordRelation.objects.create(
                            upload_record=upload_record,
                            content_type='green_bean',
                            object_id=record.id
                        )
                        
                    except Exception as e:
                        print(f"處理第 {index + 1} 行時發生錯誤: {str(e)}")
                        continue
                
                print(f"總共處理了 {len(df)} 行，跳過了 {skipped_rows} 行，成功創建了 {len(created_records)} 筆記錄")
                
                # 更新上傳記錄狀態和創建的記錄ID
                upload_record.status = 'success'
                upload_record.records_count = len(created_records)
                upload_record.created_record_ids = [str(record.id) for record in created_records]
                upload_record.save()
                
                # 記錄用戶活動
                log_user_activity(
                    request.user,
                    'upload',
                    f'上傳生豆入庫記錄檔案: {uploaded_file.name}',
                    request=request,
                    details={'records_count': len(created_records)}
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'檔案上傳成功！共處理了 {len(created_records)} 筆記錄',
                    'records_count': len(created_records)
                })
                
            except Exception as e:
                upload_record.status = 'failed'
                upload_record.error_message = str(e)
                upload_record.save()
                return JsonResponse({
                    'success': False,
                    'message': f'檔案處理失敗: {str(e)}'
                })
                
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'上傳過程中發生錯誤: {str(e)}'
        })


def calculate_file_hash(file):
    """計算檔案的MD5雜湊值"""
    hash_md5 = hashlib.md5()
    current_position = file.tell()
    file.seek(0)
    
    for chunk in file.chunks():
        hash_md5.update(chunk)
    
    file.seek(current_position)
    return hash_md5.hexdigest()


@require_green_bean_permission('delete')
@csrf_exempt
@require_http_methods(["POST"])
def delete_green_bean_record(request, record_id):
    """刪除生豆入庫記錄"""
    try:
        record = get_object_or_404(GreenBeanInboundRecord, id=record_id)
        
        # 保存記錄資訊用於記錄
        record_info = {
            'order_number': record.order_number,
            'green_bean_name': record.green_bean_name,
            'green_bean_code': record.green_bean_code,
            'green_bean_batch_number': record.green_bean_batch_number,
            'required_weight_kg': float(record.required_weight_kg) if record.required_weight_kg else None,
            'measured_weight_kg': float(record.measured_weight_kg) if record.measured_weight_kg else None,
            'execution_status': record.execution_status,
            'is_abnormal': record.is_abnormal,
        }
        
        # 記錄刪除活動（在刪除前記錄）
        log_user_activity(
            request.user,
            'delete',
            f'刪除生豆入庫記錄: {record.order_number} - {record.green_bean_name}',
            request=request,
            details={
                'record_id': str(record_id),
                'deleted_record': record_info,
                'deletion_time': datetime.now().isoformat()
            }
        )
        
        # 刪除記錄
        record.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'記錄 {record_info["order_number"]} 已成功刪除'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'刪除記錄時發生錯誤: {str(e)}'
        })


@require_green_bean_permission('delete')
@csrf_exempt
@require_http_methods(["POST"])
def batch_delete_green_bean_records(request):
    """批量刪除生豆入庫記錄"""
    try:
        data = json.loads(request.body)
        record_ids = data.get('record_ids', [])
        
        if not record_ids:
            return JsonResponse({
                'success': False,
                'message': '請選擇要刪除的記錄'
            })
        
        # 獲取要刪除的記錄
        records = GreenBeanInboundRecord.objects.filter(id__in=record_ids)
        deleted_count = records.count()
        
        # 記錄批量刪除活動
        log_user_activity(
            request.user,
            'batch_delete',
            f'批量刪除生豆入庫記錄: {deleted_count} 筆',
            request=request,
            details={'record_ids': record_ids}
        )
        
        # 刪除記錄
        records.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'已成功刪除 {deleted_count} 筆記錄'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'批量刪除記錄時發生錯誤: {str(e)}'
        })


@require_green_bean_permission('delete')
@require_http_methods(["DELETE"])
def delete_upload_record(request, upload_id):
    """刪除上傳記錄及相關的資料庫記錄"""
    try:
        # 獲取上傳記錄
        upload_record = get_object_or_404(FileUploadRecord, id=upload_id)
        
        # 檢查權限（可選：只有上傳者或管理員可以刪除）
        if upload_record.uploaded_by != request.user and not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'message': '您沒有權限刪除此上傳記錄'
            }, status=403)
        
        # 使用事務確保資料一致性
        with transaction.atomic():
            deleted_count = 0
            deleted_record_ids = []
            relation_count = 0
            
            # 方法1：通過 UploadRecordRelation 查找並刪除相關記錄
            relations = UploadRecordRelation.objects.filter(upload_record=upload_record)
            relation_count = relations.count()
            
            for relation in relations:
                try:
                    # 根據 content_type 找到對應的記錄並刪除
                    if relation.content_type == 'green_bean':
                        record = GreenBeanInboundRecord.objects.get(id=relation.object_id)
                        deleted_record_ids.append(str(record.id))
                        record.delete()
                        deleted_count += 1
                except GreenBeanInboundRecord.DoesNotExist:
                    # 記錄可能已經被刪除了，繼續處理
                    pass
            
            # 刪除關聯記錄
            relations.delete()
            
            # 方法2：如果還有 created_record_ids 中的記錄，也一併刪除
            if upload_record.created_record_ids:
                for record_id in upload_record.created_record_ids:
                    if str(record_id) not in deleted_record_ids:  # 避免重複刪除
                        try:
                            record = GreenBeanInboundRecord.objects.get(id=record_id)
                            record.delete()
                            deleted_count += 1
                            deleted_record_ids.append(str(record_id))
                        except GreenBeanInboundRecord.DoesNotExist:
                            # 記錄可能已經被刪除了，繼續處理
                            pass
            
            # 方法3：清理所有與此上傳記錄相關的孤立記錄
            # 檢查是否還有其他上傳記錄引用這些記錄
            other_upload_records = FileUploadRecord.objects.exclude(id=upload_record.id)
            protected_record_ids = set()
            
            for other_upload in other_upload_records:
                if other_upload.created_record_ids:
                    protected_record_ids.update([str(rid) for rid in other_upload.created_record_ids])
            
            # 找出不被其他上傳記錄保護的記錄並刪除
            records_to_delete = [rid for rid in deleted_record_ids if rid not in protected_record_ids]
            
            # 清理可能存在的其他關聯記錄
            other_relations = UploadRecordRelation.objects.filter(
                content_type='green_bean',
                object_id__in=records_to_delete
            ).exclude(upload_record=upload_record)
            
            additional_relation_count = other_relations.count()
            if other_relations.exists():
                other_relations.delete()
                relation_count += additional_relation_count
            
            # 最後安全檢查：如果沒有其他上傳記錄，且儀表板仍顯示資料，則清理所有孤立記錄
            remaining_upload_count = FileUploadRecord.objects.exclude(id=upload_record.id).count()
            if remaining_upload_count == 0:
                # 這是最後一個上傳記錄，清理所有可能的孤立記錄
                orphaned_records = GreenBeanInboundRecord.objects.all()
                additional_deleted = orphaned_records.count()
                if additional_deleted > 0:
                    orphaned_records.delete()
                    deleted_count += additional_deleted
                    
                # 清理所有關聯記錄
                all_relations = UploadRecordRelation.objects.filter(content_type='green_bean')
                additional_relations = all_relations.count()
                if additional_relations > 0:
                    all_relations.delete()
                    relation_count += additional_relations
            
            # 記錄用戶活動
            log_user_activity(
                request.user,
                'delete_upload_record',
                f'刪除上傳記錄 {upload_record.file_name}，刪除了 {deleted_count} 筆生豆記錄，清理了 {relation_count} 筆關聯記錄',
                request=request,
                details={
                    'upload_id': str(upload_record.id),
                    'deleted_records': deleted_record_ids,
                    'deleted_count': deleted_count,
                    'relation_count': relation_count,
                    'is_last_upload': remaining_upload_count == 0
                }
            )
            
            # 刪除上傳記錄
            file_name = upload_record.file_name
            upload_record.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'成功刪除上傳記錄 "{file_name}" 及 {deleted_count} 筆相關資料，清理了 {relation_count} 筆關聯記錄',
                'deleted_count': deleted_count,
                'relation_count': relation_count
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'刪除上傳記錄時發生錯誤: {str(e)}'
        }, status=500)


@login_required
def get_upload_records(request):
    """獲取上傳記錄列表（用於AJAX刷新）"""
    try:
        # 獲取所有生豆上傳記錄，不限制數量
        recent_uploads = FileUploadRecord.objects.filter(
            file_type='green_bean'
        ).order_by('-upload_time')
        
        upload_data = []
        for upload in recent_uploads:
            upload_data.append({
                'id': str(upload.id),
                'file_name': upload.file_name,
                'file_size': upload.file_size,
                'file_hash': upload.file_hash,
                'upload_time': upload.upload_time.strftime('%Y-%m-%d %H:%M'),
                'uploaded_by': upload.uploaded_by.username if upload.uploaded_by else '未知',
                'records_count': upload.records_count or 0,
                'status': upload.status,
                'status_display': upload.get_status_display(),
                'error_message': upload.error_message,
                'can_delete': upload.uploaded_by == request.user or request.user.is_superuser
            })
        
        return JsonResponse({
            'success': True,
            'uploads': upload_data,
            'total_count': len(upload_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'獲取上傳記錄時發生錯誤: {str(e)}'
        }, status=500)


@login_required
@login_required
@require_http_methods(["GET"])
def get_user_activities(request):
    """獲取使用者活動記錄（AJAX）"""
    try:
        # 獲取篩選參數
        user_id = request.GET.get('user')
        action_type = request.GET.get('type')
        date_range = request.GET.get('date')
        page = int(request.GET.get('page', 1))
        
        # 基礎查詢
        activities = UserActivityLog.objects.filter(
            content_type__model='greenbeaninboundrecord'
        ).select_related('user')
        
        # 應用篩選
        if user_id:
            activities = activities.filter(user_id=user_id)
        
        if action_type:
            activities = activities.filter(action=action_type)
        
        if date_range:
            now = datetime.now()
            if date_range == 'today':
                start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
                activities = activities.filter(created_at__gte=start_date)
            elif date_range == 'week':
                start_date = now - timedelta(days=7)
                activities = activities.filter(created_at__gte=start_date)
            elif date_range == 'month':
                start_date = now - timedelta(days=30)
                activities = activities.filter(created_at__gte=start_date)
        
        # 排序和分頁
        activities = activities.order_by('-created_at')
        paginator = Paginator(activities, 10)
        page_obj = paginator.get_page(page)
        
        # 生成HTML
        activities_html = []
        for activity in page_obj:
            icon_map = {
                'create': 'fa-plus',
                'update': 'fa-edit', 
                'delete': 'fa-trash',
                'upload': 'fa-upload',
                'batch_delete': 'fa-trash-alt',
                'delete_upload_record': 'fa-file-excel'
            }
            
            title_map = {
                'create': '新增記錄',
                'update': '編輯記錄',
                'delete': '刪除記錄', 
                'upload': '上傳檔案',
                'batch_delete': '批量刪除',
                'delete_upload_record': '刪除檔案'
            }
            
            icon = icon_map.get(activity.action, 'fa-info')
            title = title_map.get(activity.action, activity.action.title())
            
            # 根據不同的活動類型調整標題
            if activity.action == 'update' and activity.details:
                if activity.details.get('update_source') == 'admin_backend':
                    title = '後台編輯記錄'
                else:
                    title = '編輯記錄'
            elif activity.action == 'delete' and activity.details:
                if activity.details.get('deletion_source') == 'admin_backend':
                    title = '後台刪除記錄'
                else:
                    title = '刪除記錄'
            elif activity.action == 'create' and activity.details:
                if activity.details.get('creation_source') == 'admin_backend':
                    title = '後台新增記錄'
                else:
                    title = '新增記錄'
            
            # 生成額外的詳細資訊
            detail_info = ""
            if activity.details:
                if activity.action == 'update' and activity.details.get('changed_fields'):
                    changed_count = len(activity.details.get('changed_fields', []))
                    detail_info = f'<div class="text-muted mt-1"><i class="fas fa-edit me-1"></i>變更了 {changed_count} 個欄位</div>'
                elif activity.action in ['delete', 'batch_delete'] and activity.details.get('records_count'):
                    count = activity.details.get('records_count', 1)
                    detail_info = f'<div class="text-muted mt-1"><i class="fas fa-database me-1"></i>刪除了 {count} 筆記錄</div>'
                elif activity.action == 'upload' and activity.details.get('records_count'):
                    count = activity.details.get('records_count', 0)
                    detail_info = f'<div class="text-muted mt-1"><i class="fas fa-database me-1"></i>上傳了 {count} 筆記錄</div>'
            
            activity_html = f'''
            <div class="activity-card {activity.action}">
                <div class="d-flex">
                    <div class="activity-icon {activity.action}">
                        <i class="fas {icon}"></i>
                    </div>
                    <div class="activity-content">
                        <div class="activity-title">{title}</div>
                        <div class="activity-description">{activity.description or "執行了相關操作"}</div>
                        {detail_info}
                        <div class="activity-meta">
                            <i class="fas fa-user me-1"></i>
                            {activity.user.get_full_name() or activity.user.username}
                            <i class="fas fa-clock ms-3 me-1"></i>
                            {activity.created_at.strftime("%Y-%m-%d %H:%M:%S")}
                            {f'<i class="fas fa-globe ms-3 me-1"></i>{activity.ip_address}' if activity.ip_address else ''}
                        </div>
                    </div>
                </div>
            </div>
            '''
            activities_html.append(activity_html)
        
        if not activities_html:
            activities_html = ['''
                <div class="text-center py-5">
                    <i class="fas fa-clock fa-3x text-muted mb-3"></i>
                    <p class="text-muted">沒有找到符合條件的活動記錄</p>
                </div>
            ''']
        
        return JsonResponse({
            'success': True,
            'html': ''.join(activities_html),
            'has_more': page_obj.has_next(),
            'current_page': page,
            'total_pages': paginator.num_pages
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'載入活動記錄時發生錯誤: {str(e)}'
        })


@login_required
@permission_required('app.view_useractivitylog', raise_exception=True)
def activity_log_view(request):
    """活動記錄頁面視圖"""
    # 獲取查詢參數
    user_filter = request.GET.get('user')
    action_filter = request.GET.get('action')
    
    # 基本查詢 - 使用 select_related 優化查詢
    activities = UserActivityLog.objects.filter(
        action__in=['create', 'update', 'upload', 'delete', 'delete_upload_record', 'batch_delete', 'export']
    ).select_related('user').order_by('-created_at')
    
    # 應用過濾器
    if user_filter:
        activities = activities.filter(user__username__icontains=user_filter)
    
    if action_filter:
        activities = activities.filter(action=action_filter)
    
    # 分頁處理
    paginator = Paginator(activities, 50)  # 每頁50條記錄
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 獲取所有操作類型用於篩選
    action_choices = [
        ('create', '新增'),
        ('update', '更新'),
        ('delete', '刪除'),
        ('batch_delete', '批量刪除'),
        ('upload', '上傳'),
        ('delete_upload_record', '刪除上傳記錄'),
        ('export', '匯出'),
    ]
    
    context = {
        'activities': page_obj,
        'page_obj': page_obj,
        'action_choices': action_choices,
        'current_user_filter': user_filter,
        'current_action_filter': action_filter,
    }
    
    return render(request, 'erp/activity_log.html', context)


@require_green_bean_permission('add')
@csrf_exempt
@require_http_methods(["POST"])
def add_green_bean_record(request):
    """手動新增生豆入庫記錄"""
    try:
        # 獲取表單資料 - 基本資訊
        order_number = request.POST.get('order_number', '').strip()
        roasting_sequence = request.POST.get('roasting_sequence', '').strip()
        bean_sequence = request.POST.get('bean_sequence', '').strip()
        wave = request.POST.get('wave', '').strip()
        execution_status = request.POST.get('execution_status', '').strip()
        record_date = request.POST.get('record_date')
        
        # 生豆資訊
        bean_batch = request.POST.get('bean_batch', '').strip()
        bean_type = request.POST.get('bean_type', '').strip()
        bean_name = request.POST.get('bean_name', '').strip()
        bean_inbound_description = request.POST.get('bean_inbound_description', '').strip()
        
        # 重量與數量
        bag_weight = request.POST.get('bag_weight')
        bag_count = request.POST.get('bag_count')
        request_weight = request.POST.get('request_weight')
        actual_weight = request.POST.get('actual_weight')
        
        # 狀態與備註
        status = request.POST.get('status', '').strip()
        remarks = request.POST.get('remarks', '').strip()
        
        # 驗證必填欄位
        required_fields = [order_number, roasting_sequence, bean_sequence, wave, execution_status, 
                          bean_batch, bean_type, bean_name, bag_weight, bag_count, 
                          request_weight, actual_weight, record_date, status]
        
        if not all(field for field in required_fields if field != ''):
            return JsonResponse({'success': False, 'message': '請填写所有必填欄位'})
        
        # 轉換數據類型
        try:
            roasting_sequence = int(roasting_sequence)
            bean_sequence = int(bean_sequence)
            wave = int(wave)
            bag_weight = float(bag_weight)
            bag_count = int(bag_count)
            request_weight = float(request_weight)
            actual_weight = float(actual_weight)
            record_date = datetime.strptime(record_date, '%Y-%m-%dT%H:%M')
        except (ValueError, TypeError) as e:
            return JsonResponse({'success': False, 'message': '資料格式錯誤，請檢查數值和時間格式'})
        
        # 驗證數值
        if bag_weight <= 0 or bag_count <= 0 or request_weight <= 0 or actual_weight <= 0:
            return JsonResponse({'success': False, 'message': '重量和數量必須大於0'})
        
        if roasting_sequence <= 0 or bean_sequence <= 0 or wave <= 0:
            return JsonResponse({'success': False, 'message': '項次和波次必須大於0'})
        
        # 檢查訂單編號是否已存在
        existing_record = GreenBeanInboundRecord.objects.filter(order_number=order_number).first()
        if existing_record:
            return JsonResponse({'success': False, 'message': f'訂單編號 {order_number} 已存在，請使用不同的訂單編號'})
        
        # 設定狀態對應
        is_abnormal = (status == '異常')
        
        # 建立新記錄
        new_record = GreenBeanInboundRecord.objects.create(
            # 基本資訊
            order_number=order_number,
            roasted_item_sequence=roasting_sequence,
            green_bean_item_sequence=bean_sequence,
            batch_sequence=wave,
            execution_status=execution_status,
            record_time=record_date,
            is_abnormal=is_abnormal,
            
            # 生豆資訊
            green_bean_batch_number=bean_batch,
            green_bean_code=bean_type,
            green_bean_name=bean_name,
            green_bean_storage_silo=bean_inbound_description,  # 使用筒倉欄位存儲簡含
            
            # 重量和數量
            bag_weight_kg=bag_weight,
            input_bag_count=bag_count,
            required_weight_kg=request_weight,
            measured_weight_kg=actual_weight,
            
            # 備註
            remark=remarks
        )
        
        # 記錄用戶活動
        log_user_activity(
            request.user,
            'create',
            f'手動新增生豆入庫記錄: {order_number}',
            content_object=new_record,
            request=request,
            details={
                'order_number': order_number,
                'roasted_item_sequence': roasting_sequence,
                'green_bean_item_sequence': bean_sequence,
                'batch_sequence': wave,
                'execution_status': execution_status,
                'green_bean_name': bean_name,
                'green_bean_code': bean_type,
                'green_bean_batch_number': bean_batch,
                'bag_weight_kg': float(bag_weight),
                'input_bag_count': bag_count,
                'required_weight_kg': float(request_weight),
                'measured_weight_kg': float(actual_weight),
                'status': status
            }
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'成功新增生豆入庫記錄 {order_number}',
            'record_id': str(new_record.id)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'新增記錄時發生錯誤: {str(e)}'
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def add_activity_record(request):
    """手動新增活動記錄"""
    try:
        action = request.POST.get('action')
        description = request.POST.get('description')
        record_time = request.POST.get('record_time')
        ip_address = request.POST.get('ip_address')
        details = request.POST.get('details')
        
        # 驗證必填欄位
        if not action or not description:
            return JsonResponse({'success': False, 'message': '請填寫操作類型和操作描述'})
        
        # 轉換時間格式（如果提供的話）
        record_datetime = None
        if record_time:
            try:
                from datetime import datetime
                record_datetime = datetime.strptime(record_time, '%Y-%m-%dT%H:%M')
            except (ValueError, TypeError) as e:
                return JsonResponse({'success': False, 'message': '時間格式錯誤'})
        
        # 處理詳細資訊（JSON）
        details_dict = {}
        if details and details.strip():
            try:
                import json
                details_dict = json.loads(details)
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': '詳細資訊必須是有效的JSON格式'})
        
        # 如果沒有提供IP地址，從請求中獲取
        if not ip_address:
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0].strip()
            else:
                ip_address = request.META.get('REMOTE_ADDR')
        
        # 建立新的活動記錄（created_at 會自動設定為當前時間）
        activity_record = UserActivityLog.objects.create(
            user=request.user,
            action=action,
            description=description,
            ip_address=ip_address,
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            details=details_dict
        )
        
        # 如果用戶設定了特定時間，我們手動更新
        if record_datetime:
            activity_record.created_at = record_datetime
            activity_record.save(update_fields=['created_at'])
        
        # 同時記錄到 Django 管理員日誌系統
        _log_to_django_admin(request.user, action, description)
        
        return JsonResponse({
            'success': True,
            'message': f'成功新增活動記錄: {description}',
            'record_id': str(activity_record.id)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'新增記錄時發生錯誤: {str(e)}'
        })


@api_view(['GET'])
@require_green_bean_permission('view')
def green_bean_names_api(request):
    """獲取生豆名稱列表API"""
    try:
        from app.utils.green_bean_utils import get_green_bean_names
        green_bean_names = get_green_bean_names()
        
        return Response({
            'success': True,
            'data': green_bean_names,
            'count': len(green_bean_names)
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@require_raw_material_permission('add')
def raw_material_upload_page(request):
    """原料倉管理上傳頁面"""
    return render(request, 'erp/raw_material_import.html', {
        'title': '原料倉管理上傳',
        'upload_url': '/erp/raw-material-records/upload-file/',
        'records_url': '/erp/raw-material-records/uploads/'
    })


@require_raw_material_permission('add')
@csrf_exempt
@require_http_methods(["POST"])
def raw_material_upload_file(request):
    """原料倉管理檔案上傳處理"""
    try:
        uploaded_file = request.FILES.get('file')
        
        if not uploaded_file:
            return JsonResponse({'success': False, 'message': '請選擇要上傳的檔案'})
        
        # 檢查檔案格式
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': '只支援 .xlsx 和 .xls 格式的檔案'})
        
        # 計算檔案雜湊值（在事務外）
        file_hash = calculate_file_hash(uploaded_file)
        
        # 檢查是否為重複檔案（在事務外）
        existing_file = FileUploadRecord.objects.filter(file_hash=file_hash).first()
        if existing_file:
            return JsonResponse({
                'success': False, 
                'message': f'此檔案已於 {existing_file.upload_time.strftime("%Y-%m-%d %H:%M")} 上傳過',
                'duplicate': True
            })
        
        # 將檔案內容讀取到記憶體中（在事務外）
        file_content = uploaded_file.read()
        uploaded_file.seek(0)  # 重置檔案指針
        
        # 使用事務處理整個上傳過程
        with transaction.atomic():
            # 創建上傳記錄
            upload_record = FileUploadRecord.objects.create(
                file_name=uploaded_file.name,
                file_hash=file_hash,
                file_size=uploaded_file.size,
                uploaded_by=request.user,
                file_type='raw_material',
                status='pending'
            )
            
            # 使用 openpyxl 處理 Excel 檔案（與 test_excel_to_json.py 相同的邏輯）
            from openpyxl import load_workbook
            
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
            
            # 展開合併的儲存格
            def expand_merged_cells(ws):
                merged_ranges = list(ws.merged_cells.ranges)
                for merged_range in merged_ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    value = ws.cell(row=min_row, column=min_col).value
                    ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            ws.cell(row=row, column=col).value = value
            
            expand_merged_cells(ws)
            
            # 從檔案名稱提取月份資訊（與 test_excel_to_json.py 相同）
            def extract_month_from_filename(filename: str) -> int:
                """從檔案名稱提取月份資訊"""
                # 匹配模式：原料倉進出a2023-11.xlsx 中的 11
                pattern = r'原料倉進出a\d{4}-(\d{1,2})'
                match = re.search(pattern, filename)
                if match:
                    return int(match.group(1))
                else:
                    # 如果沒有匹配到，嘗試其他模式
                    pattern2 = r'-(\d{1,2})\.'
                    match2 = re.search(pattern2, filename)
                    if match2:
                        return int(match2.group(1))
                    else:
                        raise ValueError(f"無法從檔案名稱 {filename} 提取月份資訊")
            
            # 自動尋找標題列和子標題列
            def find_header_rows(ws) -> tuple[int, int]:
                header_row = None
                sub_header_row = None
                
                for row_num in range(1, 15):  # 檢查前15列
                    row_values = [cell.value for cell in ws[row_num]]
                    row_str = ' '.join(str(v) for v in row_values if v is not None)
                    
                    # 尋找主標題列（包含品號、品名）
                    if '品號' in row_str and '品名' in row_str and header_row is None:
                        header_row = row_num
                        
                    # 尋找子標題列（包含入庫、領用、轉出）
                    if '入庫' in row_str and '領用' in row_str and '轉出' in row_str:
                        sub_header_row = row_num
                        
                    # 如果找到主標題列，檢查下一列是否為子標題
                    if header_row is not None and sub_header_row is None:
                        next_row_values = [cell.value for cell in ws[header_row + 1]]
                        next_row_str = ' '.join(str(v) for v in next_row_values if v is not None)
                        if '入庫' in next_row_str or '領用' in next_row_str or '轉出' in next_row_str:
                            sub_header_row = header_row + 1
                
                if header_row is None:
                    header_row = 2  # 預設第二列
                if sub_header_row is None:
                    sub_header_row = header_row + 1  # 預設主標題列後一列
                    
                return header_row, sub_header_row
            
            def find_data_start_row(ws, sub_header_row: int) -> int:
                for row_num in range(sub_header_row + 1, sub_header_row + 10):
                    row_values = [cell.value for cell in ws[row_num]]
                    if any(v is not None for v in row_values):
                        return row_num
                return sub_header_row + 2
            
            def clean_column_names(columns: list[Any]) -> list[str]:
                cleaned = []
                for col in columns:
                    if col is None:
                        cleaned.append(None)
                    else:
                        cleaned_name = str(col).strip().replace('\n', ' ').replace('\r', ' ')
                        cleaned_name = ' '.join(cleaned_name.split())
                        cleaned.append(cleaned_name)
                return cleaned
            
            def merge_headers(main_headers: list[str], sub_headers: list[str]) -> list[str]:
                """合併主標題和子標題"""
                merged_headers = []
                seen_fields = set()  # 追蹤已見過的欄位名稱
                
                for i, (main_header, sub_header) in enumerate(zip(main_headers, sub_headers)):
                    if main_header is None or main_header == '':
                        # 處理只有子標題的情況
                        if sub_header is None or sub_header == '':
                            merged_headers.append(None)
                        else:
                            field_name = sub_header
                            # 檢查是否應該是小計的子欄位
                            if sub_header in ['入庫', '轉出'] and i > 0:
                                # 檢查前面是否有小計欄位
                                prev_main = main_headers[i-1] if i > 0 else None
                                if prev_main == '小計':
                                    field_name = f"小計_{sub_header}"
                                elif prev_main == '盤盈虧(外賣)':
                                    # 如果前面是盤盈虧(外賣)，且當前是入庫或轉出，則視為小計的子欄位
                                    field_name = f"小計_{sub_header}"
                            
                            if field_name in seen_fields:
                                field_name = f"{field_name}_after"
                            seen_fields.add(field_name)
                            merged_headers.append(field_name)
                    elif sub_header is None or sub_header == '':
                        # 處理只有主標題的情況
                        field_name = main_header
                        # 過濾掉不存在的欄位
                        if field_name in ['待處理', '外賣', '盤盈虧(外賣)']:
                            field_name = None
                        else:
                            if field_name in seen_fields:
                                field_name = f"{main_header}_after"
                            seen_fields.add(field_name)
                        merged_headers.append(field_name)
                    else:
                        # 處理主標題和子標題都存在的情況
                        if any(char.isdigit() for char in str(main_header)) and '/' in str(main_header):
                            # 日期格式的主標題
                            field_name = f"{main_header}_{sub_header}"
                        elif main_header == '盤盈虧(外賣)':
                            # 盤盈虧(外賣) 下的子欄位
                            field_name = f"{main_header}_{sub_header}"
                        elif main_header == '小計':
                            # 小計欄位下的子欄位
                            field_name = f"{main_header}_{sub_header}"
                        elif main_header == '領用' and sub_header == '小計':
                            # 領用_小計 特殊欄位
                            field_name = f"{main_header}_{sub_header}"
                        elif sub_header in ['入庫', '領用', '轉出'] and main_header == '小計':
                            # 小計下的入庫、領用、轉出子欄位
                            field_name = f"{main_header}_{sub_header}"
                        else:
                            # 一般主標題
                            field_name = main_header
                            
                            # 過濾掉不存在的欄位
                            if field_name in ['待處理', '外賣']:
                                field_name = None
                        
                        # 檢查是否重複
                        if field_name in seen_fields:
                            field_name = f"{field_name}_after"
                        seen_fields.add(field_name)
                        merged_headers.append(field_name)
                
                return merged_headers
            
            def analyze_column_structure(columns: list[str], file_month: int) -> dict:
                """分析欄位結構，識別月份欄位等"""
                analysis = {
                    'month_inventory': None,  # 月份庫存欄位
                    'basic_fields': [],       # 基本欄位
                    'date_fields': [],        # 日期欄位
                    'summary_fields': [],     # 小計欄位
                    'file_month': file_month, # 檔案月份
                    'found_months': []        # 找到的所有月份欄位
                }
                
                # 預期的月份庫存欄位名稱
                expected_month_inventory = f"{file_month}月庫存"
                
                for col in columns:
                    if col is None:
                        continue
                        
                    col_str = str(col)
                    
                    # 識別所有月份庫存欄位（只匹配實際存在的格式）
                    month_match = re.search(r'(\d+)月\s*庫存', col_str)
                    if month_match:
                        found_month = int(month_match.group(1))
                        # 只處理實際存在的月份欄位，避免產生不存在的欄位
                        if col_str in [f"{found_month}月 庫存", f"{found_month}月庫存"]:
                            analysis['found_months'].append((col, found_month))
                            
                            # 如果是檔案對應的月份，設為主要月份庫存欄位
                            if found_month == file_month:
                                analysis['month_inventory'] = col
                            # 如果還沒找到主要月份欄位，使用找到的第一個
                            elif analysis['month_inventory'] is None:
                                analysis['month_inventory'] = col
                            
                    # 識別 *月**日 庫存 欄位
                    elif col_str == '*月**日 庫存':
                        analysis['month_inventory'] = col
                            
                    # 識別基本欄位
                    elif col_str in ['品號', '品名', '工廠批號', '國際批號', '公斤', '包數']:
                        analysis['basic_fields'].append(col)
                    # 識別日期欄位
                    elif re.search(r'\d+/\d+', col_str):
                        analysis['date_fields'].append(col)
                    # 識別小計欄位
                    elif '小計' in col_str:
                        analysis['summary_fields'].append(col)
                
                return analysis
            
            def is_numeric_field(field_name: str) -> bool:
                """判斷欄位是否為數值型別"""
                numeric_patterns = [
                    r'公斤$',
                    r'進貨$',
                    r'領用$',
                    r'轉出$',
                    r'入庫$',
                    r'小計$',
                    r'包數$',
                    r'盤盈虧',
                    r'^\d+/\d+',  # 日期格式的數值欄位
                    r'^\d+/\d+掛\d+/\d+帳',  # 特殊日期格式
                    r'^\*月\*\*日 庫存',  # 動態月份庫存
                    r'包數_after$',  # 包數_after
                    r'\*月\*\*日 庫存_after$',  # 動態月份庫存_after
                ]
                
                for pattern in numeric_patterns:
                    if re.search(pattern, field_name):
                        return True
                return False
            
            # 從檔案名稱提取月份
            try:
                file_month = extract_month_from_filename(uploaded_file.name)
                print(f"從檔案名稱提取的月份: {file_month}月")
            except ValueError:
                file_month = 11  # 預設值
                print(f"無法從檔案名稱提取月份，使用預設值: {file_month}月")
            
            # 自動尋找標題列和子標題列
            header_row, sub_header_row = find_header_rows(ws)
            print(f"找到主標題列: 第 {header_row} 列")
            print(f"找到子標題列: 第 {sub_header_row} 列")
            
            # 抓取主標題和子標題
            main_headers = [cell.value for cell in ws[header_row]]
            sub_headers = [cell.value for cell in ws[sub_header_row]]
            
            # 清理欄位名稱
            main_headers = clean_column_names(main_headers)
            sub_headers = clean_column_names(sub_headers)
            
            print(f"主標題: {main_headers}")
            print(f"子標題: {sub_headers}")
            
            # 合併標題
            all_columns = merge_headers(main_headers, sub_headers)
            print(f"合併後欄位名稱: {all_columns}")
            
            # 分析欄位結構（與 test_excel_to_json.py 相同）
            column_analysis = analyze_column_structure(all_columns, file_month)
            print(f"\n欄位結構分析:")
            print(f"檔案月份: {column_analysis['file_month']}月")
            print(f"主要月份庫存欄位: {column_analysis['month_inventory']}")
            print(f"找到的所有月份欄位: {column_analysis['found_months']}")
            print(f"基本欄位: {column_analysis['basic_fields']}")
            print(f"日期欄位數量: {len(column_analysis['date_fields'])}")
            print(f"小計欄位: {column_analysis['summary_fields']}")
            
            # 尋找資料開始列
            data_start_row = find_data_start_row(ws, sub_header_row)
            print(f"資料開始列: 第 {data_start_row} 列")
            
            # 定義helper函數（與生豆入庫相同）
            def safe_numeric(value, default=None):
                if value is None or str(value).strip() in ['', 'nan', 'None']:
                    return default
                try:
                    result = float(value)
                    if math.isnan(result):
                        return default
                    return result
                except:
                    return default
            
            def safe_string(value, default=''):
                if value is None or str(value).strip() in ['nan', 'None']:
                    return default
                return str(value).strip()
            
            def safe_string_from_numeric(value, default=''):
                """處理可能是數字的字符串欄位"""
                if value is None or str(value).strip() in ['nan', 'None']:
                    return default
                try:
                    if isinstance(value, (int, float)):
                        return str(int(value))
                    return str(value).strip()
                except:
                    return str(value).strip()
            
            def safe_integer(value, default=None):
                """安全轉換為整數"""
                if value is None or str(value).strip() in ['', 'nan', 'None']:
                    return default
                try:
                    result = float(value)
                    if math.isnan(result):
                        return default
                    return int(result)
                except:
                    return default
            
            # 處理資料 - 完全使用 test_excel_to_json.py 的邏輯
            created_records = []
            skipped_rows = 0
            row_count = 0
            
            # 定義 RawMaterialRow.from_row 方法（與 test_excel_to_json.py 完全相同）
            def from_row(row: list[Any], columns: list[str]) -> dict:
                data = {}
                for col_name, value in zip(columns, row):
                    if col_name is None:
                        continue
                    col_name = str(col_name).strip().replace('\n', ' ')  # 清理換行符號
                    key = '公斤' if col_name == '標準重' else col_name
                    # 自動型別轉換
                    if key == '包數':
                        try:
                            data[key] = math.ceil(float(value)) if value is not None else None
                        except (TypeError, ValueError):
                            data[key] = None
                    elif is_numeric_field(key):
                        try:
                            data[key] = float(value) if value is not None else None
                        except (TypeError, ValueError):
                            data[key] = None
                    else:
                        data[key] = str(value) if value is not None else None
                return data
            
            for row in ws.iter_rows(min_row=data_start_row):
                row_count += 1
                try:
                    values = [cell.value for cell in row]
                    # 若全為 None 則跳過
                    if all(v is None for v in values):
                        continue
                    
                    # 使用與 test_excel_to_json.py 完全相同的 from_row 方法
                    row_data = from_row(values, all_columns)
                    
                    # 只保留公斤有值的資料（與 test_excel_to_json.py 完全一致）
                    if row_data.get('公斤') is None:
                        print(f"跳過第 {row_count} 行：公斤為空")
                        skipped_rows += 1
                        continue
                    
                    # 獲取品號和品名（不檢查是否為空，與 test_excel_to_json.py 一致）
                    product_code = str(row_data.get('品號', '')).strip()
                    product_name = str(row_data.get('品名', '')).strip()
                    
                    print(f"處理第 {row_count} 行: 品號='{product_code}', 品名='{product_name}', 公斤='{row_data.get('公斤')}'")
                    
                    # 分離基本欄位和動態欄位
                    basic_fields = {
                        'product_code': product_code,
                        'product_name': product_name,
                        'factory_batch_number': str(row_data.get('工廠批號', '')) if row_data.get('工廠批號') is not None else '',
                        'international_batch_number': str(row_data.get('國際批號', '')) if row_data.get('國際批號') is not None else '',
                        'standard_weight_kg': row_data.get('公斤', 0) or 0,
                        'record_date': datetime.now().date()
                    }
                    
                    # 處理月庫存欄位（根據檔名動態變化）
                    file_month = extract_month_from_filename(uploaded_file.name)
                    file_year = int(re.search(r'(\d{4})-\d{1,2}', uploaded_file.name).group(1)) if re.search(r'(\d{4})-\d{1,2}', uploaded_file.name) else datetime.now().year
                    import calendar
                    
                    # 上月庫存為檔名月份-2
                    prev2_month = (file_month - 2) % 12 or 12
                    prev2_year = file_year if file_month > 2 else file_year - 1
                    previous_month_key = f"{prev2_month}月 庫存"
                    
                    # 調試輸出
                    print(f"檔名月份: {file_month}, 上月庫存欄位: {previous_month_key}")
                    print(f"可用的欄位: {list(row_data.keys())}")
                    
                    # 將基本欄位也放入基本欄位中
                    basic_fields.update({
                        'previous_month_inventory': row_data.get(previous_month_key, 0) or 0,  # 上月庫存
                        'incoming_stock': row_data.get('進貨', 0) or 0,  # 進貨
                        'outgoing_stock': row_data.get('領用', 0) or 0,  # 領用
                        'current_inventory': row_data.get('*月**日 庫存', 0) or 0,  # 當前庫存
                    })
                    
                    # 收集所有動態欄位（只包含上方基本欄位中沒有的內容）
                    dynamic_fields = {}
                    
                    # 排除所有基本欄位，只保留日期相關欄位和其他特殊欄位
                    basic_field_names = [
                        '品號', '品名', '工廠批號', '國際批號', '公斤', '包數',
                        '進貨', '領用', previous_month_key, '*月**日 庫存'
                    ]
                    
                    # 定義允許的動態欄位模式（更嚴格）
                    allowed_dynamic_patterns = [
                        r'^\d+/\d+掛\d+/\d+帳_',  # 10/31掛11/1帳_入庫
                        r'^\d+/\d+_',  # 11/1_入庫, 11/1_領用, 11/1_轉出
                        r'^盤盈虧\(外賣\)_',  # 盤盈虧(外賣)_入庫
                        r'^小計_',  # 小計_入庫, 小計_領用, 小計_轉出
                        r'^領用_小計$',  # 領用_小計
                        r'^\*月\*\*日 庫存_after$',  # *月**日 庫存_after
                        r'^包數_after$',  # 包數_after
                    ]
                    
                    # 明確排除的欄位
                    excluded_fields = {
                        '待處理', '外賣', '盤盈虧(外賣)',  # 這些欄位不應該存在
                    }
                    
                    # 排除所有月份庫存欄位（除了當前動態產生的）
                    for key in list(row_data.keys()):
                        if re.match(r'^\d+月\s*庫存$', key) and key != previous_month_key:
                            excluded_fields.add(key)
                    
                    # 產生正確的日期動態欄位名稱（根據檔名）
                    days_in_month = calendar.monthrange(file_year, file_month)[1]
                    correct_date_fields = {}
                    
                    # 產生本月日期欄位
                    for day in range(1, days_in_month + 1):
                        for t in ['入庫', '領用', '轉出']:
                            correct_key = f"{file_month}/{day}_{t}"
                            # 尋找對應的原始欄位（可能是任何日期的欄位）
                            for original_key, original_value in row_data.items():
                                if re.match(r'^\d+/\d+_' + t + '$', original_key):
                                    correct_date_fields[correct_key] = original_value
                                    break
                    
                    # 產生跨月欄位（前一月最後一天掛本月1日帳）
                    prev_month = (file_month - 1) % 12 or 12
                    prev_month_year = file_year if file_month > 1 else file_year - 1
                    prev_month_last_day = calendar.monthrange(prev_month_year, prev_month)[1]
                    for t in ['入庫', '領用', '轉出']:
                        correct_key = f"{prev_month}/{prev_month_last_day}掛{file_month}/1帳_{t}"
                        # 尋找對應的原始跨月欄位
                        for original_key, original_value in row_data.items():
                            if re.match(r'^\d+/\d+掛\d+/\d+帳_' + t + '$', original_key):
                                correct_date_fields[correct_key] = original_value
                                break
                    
                    for key, value in row_data.items():
                        # 跳過基本欄位和明確排除的欄位
                        if key in basic_field_names or key in excluded_fields:
                            continue
                            
                        # 檢查是否為允許的動態欄位
                        is_allowed = False
                        for pattern in allowed_dynamic_patterns:
                            if re.match(pattern, key):
                                is_allowed = True
                                break
                        
                        if is_allowed:
                            # 如果是日期相關欄位，使用正確的欄位名稱
                            if re.match(r'^\d+/\d+', key) or re.match(r'^\d+/\d+掛\d+/\d+帳_', key):
                                # 日期欄位已經在 correct_date_fields 中處理
                                continue
                            else:
                                # 非日期欄位，直接使用
                                if is_numeric_field(key):
                                    dynamic_fields[key] = float(value) if value is not None else None
                                else:
                                    dynamic_fields[key] = str(value) if value is not None else None
                    
                    # 將正確的日期欄位加入動態欄位
                    for correct_key, value in correct_date_fields.items():
                        if is_numeric_field(correct_key):
                            dynamic_fields[correct_key] = float(value) if value is not None else None
                        else:
                            dynamic_fields[correct_key] = str(value) if value is not None else None
                    
                    # 調試輸出
                    print(f"動態欄位數量: {len(dynamic_fields)}")
                    if dynamic_fields:
                        print(f"動態欄位範例: {list(dynamic_fields.items())[:3]}")
                    
                    # 建立記錄（包含動態欄位）
                    record = RawMaterialWarehouseRecord.objects.create(
                        **basic_fields,
                        dynamic_fields=dynamic_fields
                    )
                    
                    created_records.append(record)
                    print(f"成功創建記錄: {record.product_code} - {record.product_name}")
                    
                    # 創建關聯記錄
                    UploadRecordRelation.objects.create(
                        upload_record=upload_record,
                        content_type='raw_material',
                        object_id=record.id
                    )
                    
                except Exception as e:
                    print(f"處理第 {row_count} 行時發生錯誤: {str(e)}")
                    continue
            
            print(f"總共處理了 {row_count} 行，跳過了 {skipped_rows} 行，成功創建了 {len(created_records)} 筆記錄")
            
            # 更新上傳記錄狀態
            upload_record.status = 'success'
            upload_record.records_count = len(created_records)
            upload_record.created_record_ids = [str(record.id) for record in created_records]
            upload_record.save()
        
        # 記錄用戶活動（在事務外）
        try:
            log_user_activity(
                request.user,
                'upload',
                f'上傳原料倉管理檔案: {uploaded_file.name}',
                request=request,
                details={'records_count': len(created_records)}
            )
        except Exception as e:
            print(f"記錄用戶活動失敗: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'檔案上傳成功！共處理了 {len(created_records)} 筆記錄',
            'records_count': len(created_records)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'上傳過程中發生錯誤: {str(e)}'
        })


@require_raw_material_permission('view')
def get_raw_material_upload_records(request):
    """獲取原料倉上傳記錄列表（用於AJAX刷新）"""
    try:
        # 獲取所有原料倉上傳記錄
        recent_uploads = FileUploadRecord.objects.filter(
            file_type='raw_material'
        ).order_by('-upload_time')
        
        upload_data = []
        for upload in recent_uploads:
            upload_data.append({
                'id': str(upload.id),
                'file_name': upload.file_name,
                'file_size': upload.file_size,
                'file_hash': upload.file_hash,
                'upload_time': upload.upload_time.strftime('%Y-%m-%d %H:%M'),
                'uploaded_by': upload.uploaded_by.username if upload.uploaded_by else '未知',
                'records_count': upload.records_count or 0,
                'status': upload.status,
                'status_display': upload.get_status_display(),
                'error_message': upload.error_message,
                'can_delete': upload.uploaded_by == request.user or request.user.is_superuser
            })
        
        return JsonResponse({
            'success': True,
            'uploads': upload_data,
            'total_count': len(upload_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'獲取上傳記錄時發生錯誤: {str(e)}'
        }, status=500)


@require_raw_material_permission('delete')
@require_http_methods(["POST"])
def delete_raw_material_upload_record(request, upload_id):
    """刪除原料倉上傳記錄"""
    try:
        upload_record = FileUploadRecord.objects.get(id=upload_id, file_type='raw_material')
        
        # 檢查權限
        if upload_record.uploaded_by != request.user and not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'message': '您沒有權限刪除此上傳記錄'
            }, status=403)
        
        # 保存上傳記錄資訊（在刪除前）
        file_name = upload_record.file_name
        upload_id_str = str(upload_record.id)
        
        with transaction.atomic():
            # 刪除相關的原料倉記錄
            relations = UploadRecordRelation.objects.filter(upload_record=upload_record)
            deleted_records = 0
            
            for relation in relations:
                try:
                    if relation.content_type == 'raw_material':
                        record = RawMaterialWarehouseRecord.objects.get(id=relation.object_id)
                        record.delete()
                        deleted_records += 1
                except RawMaterialWarehouseRecord.DoesNotExist:
                    pass
            
            # 刪除關聯記錄
            relations.delete()
            
            # 刪除上傳記錄
            upload_record.delete()
        
        # 記錄活動（在事務外）
        try:
            log_user_activity(
                request.user,
                'delete_upload_record',
                f'刪除原料倉上傳記錄: {file_name}',
                request=request,
                details={
                    'upload_id': upload_id_str,
                    'deleted_records': deleted_records
                }
            )
        except Exception as e:
            print(f"記錄刪除活動失敗: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'成功刪除上傳記錄，同時刪除了 {deleted_records} 筆相關記錄'
        })
            
    except FileUploadRecord.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '找不到指定的上傳記錄'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'刪除過程中發生錯誤: {str(e)}'
        }, status=500)